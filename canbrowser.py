import os
import shutil
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import time
import undetected_chromedriver as uc
import openpyxl
from datetime import datetime

class CanBrowser:
    def __init__(self, ID):
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.parent_dir = os.path.dirname(self.current_dir)
        # chrome profile PATH 
        self.profile_path = os.path.join(self.parent_dir, 'chrome-profile')
        # ChromeDriver PATH
        self.chrome_driver_path = os.path.join(self.parent_dir, 'chrome-driver', 'chromedriver.exe')
        # Chrome-win PATH
        self.chrome_path = os.path.join(self.parent_dir, 'chrome-win', 'chrome.exe')
        self.driver = None
        self.profile_id = ID

    def add_to_excel(self, profile_id, profile_name, profile_path, note):
        excel_file_path = os.path.join(self.parent_dir, 'profiles.xlsx')
        
        if not os.path.exists(excel_file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(['ID', 'Profile Name', 'Profile Path', 'Note', "Last Used"])
        else:
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook.active
            
        # Kiểm tra nếu ID hồ sơ đã tồn tại trong Excel
        for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=worksheet.max_row, values_only=True):
            if row[0] == profile_id:
                return  # ID đã tồn tại, không cần thêm
        
        # Tìm dòng chính xác để thêm thông tin mới
        insert_row = 2
        for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=worksheet.max_row, values_only=True):
            if row[0] > profile_id:
                break
            insert_row += 1
        
        # Chèn thông tin mới vào dòng đúng
        worksheet.insert_rows(insert_row)
        worksheet.cell(row=insert_row, column=1, value=profile_id)
        worksheet.cell(row=insert_row, column=2, value=profile_name)
        worksheet.cell(row=insert_row, column=3, value=profile_path)
        worksheet.cell(row=insert_row, column=4, value=note)
        
        workbook.save(excel_file_path)

    def update_last_opened_time(self, row_idx, last_opened_time):
        excel_file_path = os.path.join(self.parent_dir, 'profiles.xlsx')
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        worksheet.cell(row=row_idx + 2, column=5, value=last_opened_time)  # Column index is 1-based
        workbook.save(excel_file_path)

    def create_webdriver(self, profile_name):
        chrome_options = Options()
        chrome_options.binary_location = self.chrome_path
        profile_dir = os.path.join(self.profile_path, profile_name)
        chrome_options.add_argument("--user-data-dir=" + profile_dir)
        service = Service(self.chrome_driver_path)
        self.driver = uc.Chrome(service=service, options=chrome_options)
    
    def create_profile(self):
        profile_name = f"profile_{self.profile_id}"
        profile_dir = os.path.join(self.profile_path, profile_name)
        os.makedirs(profile_dir, exist_ok=True)
        self.create_webdriver(profile_name)

        self.add_to_excel(self.profile_id, profile_name, profile_dir, "note")
        
    def open_profile(self):
        self.create_profile()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.update_last_opened_time(self.profile_id, current_time) 
    
    def close_profile(self):
        if self.driver:
            self.driver.quit()

    def delete_profile(self):
        profile_name = f"profile_{self.profile_id}"
        profile_dir = os.path.join(self.profile_path, profile_name)
        if os.path.exists(profile_dir):
            shutil.rmtree(profile_dir)
            
        excel_file_path = os.path.join(self.parent_dir, 'profiles.xlsx')
        if os.path.exists(excel_file_path):
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # Tìm dòng chứa thông tin của profile để xóa
            delete_row = None
            for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=worksheet.max_row, values_only=True):
                if row[0] == self.profile_id:
                    delete_row = row[0]
                    break
            
            # Xóa dòng thông tin của profile nếu tìm thấy
            if delete_row is not None:
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    if row[0].value == delete_row:
                        worksheet.delete_rows(row[0].row)
                        break
                
                workbook.save(excel_file_path)

    def new_profile(self):
        excel_file_path = os.path.join(self.parent_dir, 'profiles.xlsx')
        
        if not os.path.exists(excel_file_path):
            profile_id = 0
        else:
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # Calculate the next profile ID based on the number of existing profiles
            profile_id = worksheet.max_row
            
        self.profile_id = profile_id
        self.create_profile()

if __name__ == "__main__":
    profile1 = CanBrowser(2)
    profile1.open_profile()
    time.sleep(2)
    if profile1.driver:
        profile1.driver.get("https://www.google.com")
    
    while True:
        time.sleep(1)
        
